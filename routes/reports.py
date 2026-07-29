import io
from flask import Blueprint, render_template, abort, request, send_file
from flask_login import login_required, current_user
from models.member import Member
from models.attendance import Event, Attendance
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

reports_bp = Blueprint("reports", __name__)

@reports_bp.route("/reports")
@login_required
def index():
    if getattr(current_user, 'tipo', None) != 'MASTER':
        abort(403)

    departamento_selecionado = request.args.get('departamento')

    # Consulta base filtrando por membros ativos
    query = Member.query.filter_by(ativo=True)
    if departamento_selecionado and departamento_selecionado.strip():
        query = query.filter_by(departamento=departamento_selecionado)
        
    members = query.all()
    total_ebds = Event.query.count() or 0
    
    relatorio_membros = []
    for member in members:
        presencas = Attendance.query.filter_by(member_id=member.id, presente=True).count()
        porcentagem = round((presencas / total_ebds) * 100, 1) if total_ebds > 0 else 0
        
        relatorio_membros.append({
            "member_id": member.id,
            "nome": getattr(member, 'nome', 'Sem Nome'),
            "departamento": getattr(member, 'departamento', 'Geral'),
            "foto": getattr(member, 'foto', None),
            "presencas": presencas,
            "total_ebds": total_ebds,
            "porcentagem": porcentagem
        })

    # Ordenação padronizada: 1º Presenças (desc), 2º Porcentagem (desc), 3º ID do Membro (asc)
    ranking_membros = sorted(relatorio_membros, key=lambda x: (-x["presencas"], -x["porcentagem"], x["member_id"]))

    return render_template("reports.html", 
                           relatorio_membros=ranking_membros, 
                           departamento_selecionado=departamento_selecionado)

@reports_bp.route("/reports/export_excel")
@login_required
def export_excel():
    if getattr(current_user, 'tipo', None) != 'MASTER':
        abort(403)

    departamento_selecionado = request.args.get('departamento')

    query = Member.query.filter_by(ativo=True)
    if departamento_selecionado and departamento_selecionado.strip():
        query = query.filter_by(departamento=departamento_selecionado)
        
    members = query.all()
    total_ebds = Event.query.count() or 0
    
    dados_excel = []
    for member in members:
        presencas = Attendance.query.filter_by(member_id=member.id, presente=True).count()
        # Cálculo de faltas baseado no total de eventos cadastrados
        faltas = total_ebds - presencas
        if faltas < 0:
            faltas = 0
            
        porcentagem = round((presencas / total_ebds) * 100, 1) if total_ebds > 0 else 0
        
        dados_excel.append({
            "member_id": member.id,
            "nome": getattr(member, 'nome', 'Sem Nome'),
            "departamento": getattr(member, 'departamento', 'Geral'),
            "presencas": presencas,
            "faltas": faltas,
            "total_ebds": total_ebds,
            "porcentagem": porcentagem
        })

    # Aplica a mesma ordenação padrão do sistema
    dados_excel.sort(key=lambda x: (-x["presencas"], -x["porcentagem"], x["member_id"]))

    # Criação da planilha Excel com openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Frequência"

    # Estilos visuais profissionais
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=10)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # Cabeçalho
    headers = ["Nome", "Departamento", "Presenças", "Faltas", "Total EBDs", "Assiduidade (%)"]
    ws.append(headers)

    # Formatação do Cabeçalho
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

    # Inserção dos Dados
    for item in dados_excel:
        ws.append([
            item["nome"],
            item["departamento"],
            item["presencas"],
            item["faltas"],
            item["total_ebds"],
            f"{item['porcentagem']}%"
        ])

    # Formatação das linhas de dados
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.font = data_font
            cell.border = border_thin
            if cell.column == 1 or cell.column == 2:
                cell.alignment = align_left
            else:
                cell.alignment = align_center

    # Ajuste automático de largura das colunas
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 5, 15)

    # Salva em memória para o envio
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    dept_suffix = f"_{departamento_selecionado.lower()}" if departamento_selecionado else ""
    filename = f"relatorio_frequencia{dept_suffix}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )